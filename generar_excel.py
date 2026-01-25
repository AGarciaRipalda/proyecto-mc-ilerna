from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Crear workbook
wb = Workbook()
ws = wb.active
ws.title = "Diccionario de Datos"

# Encabezados
headers = ['TABLA', 'Nombre', 'Tipo', 'Unidad', 'Valores', 'Valor por defecto', 'Restricciones', 'Descripción', 'Ejemplo']
ws.append(headers)

# Datos de todas las tablas
datos = [
    # REPARTIDOR
    ['REPARTIDOR', 'Num_Repartidor', 'INT', 'Número', '1 a 2147483647', 'AUTO_INCREMENT', 'PK, NOT NULL', 'Identificador único del repartidor', '1, 2, 3...'],
    ['REPARTIDOR', 'Nombre', 'VARCHAR(50)', 'Texto', 'Alfanumérico', '-', 'NOT NULL', 'Nombre del repartidor', 'Carlos'],
    ['REPARTIDOR', 'Apellido1', 'VARCHAR(50)', 'Texto', 'Alfanumérico', '-', 'NOT NULL', 'Primer apellido', 'García'],
    ['REPARTIDOR', 'Apellido2', 'VARCHAR(50)', 'Texto', 'Alfanumérico', 'NULL', '-', 'Segundo apellido (opcional)', 'López'],
    ['REPARTIDOR', 'DNI', 'VARCHAR(9)', 'Texto', '8 dígitos + letra', '-', 'UNIQUE, NOT NULL', 'Documento de identidad español', '12345678A'],
    ['REPARTIDOR', 'Telefono', 'VARCHAR(15)', 'Texto', 'Numérico con prefijo', 'NULL', '-', 'Teléfono de contacto', '+34 600123456'],
    ['REPARTIDOR', 'Matricula_Moto', 'VARCHAR(10)', 'Texto', 'Formato matrícula', 'NULL', '-', 'Matrícula del vehículo asignado', '1234ABC'],
    ['REPARTIDOR', 'Turno', 'CHAR(1)', 'Carácter', 'M, T, N', '-', "CHECK IN ('M','T','N'), NOT NULL", 'Turno de trabajo: Mañana, Tarde, Noche', 'M'],
    
    # INGREDIENTE (NUEVA TABLA)
    ['INGREDIENTE', 'Cod_Ingrediente', 'INT', 'Número', '1 a 2147483647', 'AUTO_INCREMENT', 'PK, NOT NULL', 'Identificador único del ingrediente', '1, 2, 3...'],
    ['INGREDIENTE', 'Nombre', 'VARCHAR(100)', 'Texto', 'Alfanumérico', '-', 'UNIQUE, NOT NULL', 'Nombre del ingrediente', 'Lechuga'],
    ['INGREDIENTE', 'Alergeno', 'BOOLEAN', 'Booleano', 'TRUE, FALSE', 'FALSE', 'NOT NULL', 'Indica si es un alérgeno', 'TRUE'],
    ['INGREDIENTE', 'Tipo_Alergeno', 'VARCHAR(50)', 'Texto', 'Gluten, Lactosa, etc.', 'NULL', '-', 'Tipo de alérgeno (NULL si no es alérgeno)', 'Lactosa'],
    
    # PRODUCTO (MODIFICADA - sin campo Ingredientes)
    ['PRODUCTO', 'Cod_Producto', 'INT', 'Número', '1 a 2147483647', 'AUTO_INCREMENT', 'PK, NOT NULL', 'Código único del producto', '1, 2, 3...'],
    ['PRODUCTO', 'Nombre', 'VARCHAR(100)', 'Texto', 'Alfanumérico', '-', 'NOT NULL', 'Nombre comercial del producto', 'Hamburguesa Doble'],
    ['PRODUCTO', 'Precio', 'DECIMAL(6,2)', 'Euros (€)', '0.01 a 9999.99', '-', 'CHECK > 0, NOT NULL', 'Precio unitario actual del producto', '5.50'],
    
    # MENU
    ['MENU', 'Cod_Menu', 'INT', 'Número', '1 a 2147483647', 'AUTO_INCREMENT', 'PK, NOT NULL', 'Código único del menú', '1, 2, 3...'],
    ['MENU', 'Nombre', 'VARCHAR(100)', 'Texto', 'Alfanumérico', '-', 'NOT NULL', 'Nombre comercial del menú', 'Menú Ahorro'],
    ['MENU', 'Descripcion', 'TEXT', 'Texto', 'Alfanumérico', 'NULL', '-', 'Descripción detallada de la oferta', 'Hamburguesa + Bebida + Patatas'],
    ['MENU', 'Precio', 'DECIMAL(6,2)', 'Euros (€)', '0.01 a 9999.99', '-', 'CHECK > 0, NOT NULL', 'Precio promocional del menú completo', '6.90'],
    
    # PEDIDO
    ['PEDIDO', 'Num_Pedido', 'INT', 'Número', '1 a 2147483647', 'AUTO_INCREMENT', 'PK, NOT NULL', 'Número correlativo único de pedido', '1, 2, 3...'],
    ['PEDIDO', 'Fecha', 'DATE', 'Fecha', 'YYYY-MM-DD', '-', 'NOT NULL', 'Fecha del pedido', '2026-01-21'],
    ['PEDIDO', 'Hora', 'TIME', 'Hora', 'HH:MM:SS', '-', 'NOT NULL', 'Hora del pedido', '14:30:00'],
    
    # PEDIDO_VENTANILLA
    ['PEDIDO_VENTANILLA', 'Num_Pedido', 'INT', 'Número', 'Referencia a PEDIDO', '-', 'PK, FK → PEDIDO, NOT NULL', 'Referencia al pedido base (relación 1:1)', '1, 2, 3...'],
    ['PEDIDO_VENTANILLA', 'Num_Ventanilla', 'INT', 'Número', '1 a 10', '-', 'NOT NULL', 'Número de ventanilla donde se atendió', '1, 2, 3'],
    
    # PEDIDO_DOMICILIO
    ['PEDIDO_DOMICILIO', 'Num_Pedido', 'INT', 'Número', 'Referencia a PEDIDO', '-', 'PK, FK → PEDIDO, NOT NULL', 'Referencia al pedido base (relación 1:1)', '1, 2, 3...'],
    ['PEDIDO_DOMICILIO', 'Telefono_Contacto', 'VARCHAR(15)', 'Texto', 'Numérico con prefijo', '-', 'NOT NULL', 'Teléfono del cliente para contacto', '+34 600123456'],
    ['PEDIDO_DOMICILIO', 'Poblacion', 'VARCHAR(100)', 'Texto', 'Alfanumérico', '-', 'NOT NULL', 'Población de entrega', 'Jerez de la Frontera'],
    ['PEDIDO_DOMICILIO', 'Direccion_Entrega', 'VARCHAR(200)', 'Texto', 'Alfanumérico', '-', 'NOT NULL', 'Dirección completa de entrega', 'Calle Mayor 123, 3ºB'],
    ['PEDIDO_DOMICILIO', 'Num_Repartidor', 'INT', 'Número', 'Referencia a REPARTIDOR', '-', 'FK → REPARTIDOR, NOT NULL', 'Repartidor asignado al pedido', '1, 2, 3...'],
    
    # PRODUCTO_INGREDIENTE (NUEVA TABLA - N:M)
    ['PRODUCTO_INGREDIENTE', 'Cod_Producto', 'INT', 'Número', 'Referencia a PRODUCTO', '-', 'PK (parte 1 de 2), FK → PRODUCTO, NOT NULL', 'Código del producto (parte de clave compuesta)', '1, 2, 3...'],
    ['PRODUCTO_INGREDIENTE', 'Cod_Ingrediente', 'INT', 'Número', 'Referencia a INGREDIENTE', '-', 'PK (parte 2 de 2), FK → INGREDIENTE, NOT NULL', 'Código del ingrediente (parte de clave compuesta)', '1, 2, 3...'],
    
    # COMPOSICION_MENU (N:M)
    ['COMPOSICION_MENU', 'Cod_Menu', 'INT', 'Número', 'Referencia a MENU', '-', 'PK (parte 1 de 2), FK → MENU, NOT NULL', 'Código del menú (parte de clave compuesta)', '1, 2, 3...'],
    ['COMPOSICION_MENU', 'Cod_Producto', 'INT', 'Número', 'Referencia a PRODUCTO', '-', 'PK (parte 2 de 2), FK → PRODUCTO, NOT NULL', 'Código del producto (parte de clave compuesta)', '1, 2, 3...'],
    ['COMPOSICION_MENU', 'Cantidad', 'INT', 'Unidades', '1 a 100', '-', 'CHECK > 0, NOT NULL', 'Unidades del producto en el menú', '1, 2, 3'],
    
    # DETALLE_PEDIDO_PRODUCTO (N:M)
    ['DETALLE_PEDIDO_PRODUCTO', 'Num_Pedido', 'INT', 'Número', 'Referencia a PEDIDO', '-', 'PK (parte 1 de 2), FK → PEDIDO, NOT NULL', 'Número de pedido (parte de clave compuesta)', '1, 2, 3...'],
    ['DETALLE_PEDIDO_PRODUCTO', 'Cod_Producto', 'INT', 'Número', 'Referencia a PRODUCTO', '-', 'PK (parte 2 de 2), FK → PRODUCTO, NOT NULL', 'Código del producto (parte de clave compuesta)', '1, 2, 3...'],
    ['DETALLE_PEDIDO_PRODUCTO', 'Cantidad', 'INT', 'Unidades', '1 a 100', '-', 'CHECK > 0, NOT NULL', 'Unidades solicitadas', '1, 2, 5'],
    ['DETALLE_PEDIDO_PRODUCTO', 'Precio_Venta', 'DECIMAL(6,2)', 'Euros (€)', '0.01 a 9999.99', '-', 'CHECK > 0, NOT NULL', 'Precio unitario en el momento de la venta', '5.50'],
    
    # DETALLE_PEDIDO_MENU (N:M)
    ['DETALLE_PEDIDO_MENU', 'Num_Pedido', 'INT', 'Número', 'Referencia a PEDIDO', '-', 'PK (parte 1 de 2), FK → PEDIDO, NOT NULL', 'Número de pedido (parte de clave compuesta)', '1, 2, 3...'],
    ['DETALLE_PEDIDO_MENU', 'Cod_Menu', 'INT', 'Número', 'Referencia a MENU', '-', 'PK (parte 2 de 2), FK → MENU, NOT NULL', 'Código del menú (parte de clave compuesta)', '1, 2, 3...'],
    ['DETALLE_PEDIDO_MENU', 'Cantidad', 'INT', 'Unidades', '1 a 100', '-', 'CHECK > 0, NOT NULL', 'Unidades solicitadas', '1, 2, 3'],
    ['DETALLE_PEDIDO_MENU', 'Precio_Venta', 'DECIMAL(6,2)', 'Euros (€)', '0.01 a 9999.99', '-', 'CHECK > 0, NOT NULL', 'Precio del menú en el momento de la venta', '6.90'],
]

# Agregar datos
for fila in datos:
    ws.append(fila)

# Formato de encabezados
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_num in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Formato de datos
tabla_actual = None
fill_tabla1 = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
fill_tabla2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
fill_pk_compuesta = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
fill_nueva_tabla = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde claro para nuevas tablas
usar_fill1 = True

for row_num in range(2, ws.max_row + 1):
    tabla_nombre = ws.cell(row=row_num, column=1).value
    restricciones = ws.cell(row=row_num, column=7).value
    
    # Alternar color por tabla
    if tabla_nombre != tabla_actual:
        tabla_actual = tabla_nombre
        usar_fill1 = not usar_fill1
    
    # Color especial para nuevas tablas (INGREDIENTE y PRODUCTO_INGREDIENTE)
    if tabla_nombre in ['INGREDIENTE', 'PRODUCTO_INGREDIENTE']:
        fill_actual = fill_nueva_tabla
    # Color especial para claves primarias compuestas
    elif restricciones and 'PK (parte' in restricciones:
        fill_actual = fill_pk_compuesta
    else:
        fill_actual = fill_tabla1 if usar_fill1 else fill_tabla2
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.border = border
        cell.fill = fill_actual
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        # Negrita para nombre de tabla
        if col_num == 1:
            if tabla_nombre in ['INGREDIENTE', 'PRODUCTO_INGREDIENTE']:
                cell.font = Font(bold=True, size=10, color="006100")  # Verde oscuro para nuevas tablas
            else:
                cell.font = Font(bold=True, size=10)
        
        # Negrita para claves primarias compuestas
        if col_num == 2 and restricciones and 'PK (parte' in restricciones:
            cell.font = Font(bold=True, size=10, color="CC0000")

# Ajustar anchos de columna
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 18
ws.column_dimensions['G'].width = 35
ws.column_dimensions['H'].width = 45
ws.column_dimensions['I'].width = 25

# Ajustar altura de filas con PKs compuestas
for row_num in range(2, ws.max_row + 1):
    restricciones = ws.cell(row=row_num, column=7).value
    if restricciones and 'PK (parte' in restricciones:
        ws.row_dimensions[row_num].height = 30

# Congelar primera fila
ws.freeze_panes = 'A2'

# Guardar archivo
wb.save('DICCIONARIO_DATOS.xlsx')
print("✅ Archivo Excel actualizado: DICCIONARIO_DATOS.xlsx")
print("📊 Total de filas:", len(datos) + 1, "(incluyendo encabezado)")
print("📋 Tablas incluidas: 11 tablas del sistema (2 nuevas)")
print("")
print("🆕 NUEVAS TABLAS:")
print("   • INGREDIENTE (4 campos) - Catálogo de ingredientes con alérgenos")
print("   • PRODUCTO_INGREDIENTE (2 campos - PK compuesta) - Relación N:M")
print("")
print("📝 MODIFICACIONES:")
print("   • PRODUCTO: Eliminado campo 'Ingredientes' (ahora normalizado)")
print("")
print("🔑 CLAVES PRIMARIAS:")
print("   • Simples: REPARTIDOR, INGREDIENTE, PRODUCTO, MENU, PEDIDO, PEDIDO_VENTANILLA, PEDIDO_DOMICILIO")
print("   • Compuestas: PRODUCTO_INGREDIENTE, COMPOSICION_MENU, DETALLE_PEDIDO_PRODUCTO, DETALLE_PEDIDO_MENU")
print("")
print("🎨 FORMATO:")
print("   • Nuevas tablas: Fondo verde claro y nombre en verde oscuro")
print("   • Claves primarias compuestas: Fondo amarillo y texto rojo")
print("   • Tablas existentes: Gris claro / Blanco alternado")
